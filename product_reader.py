import os
import openpyxl
from datetime import datetime
from product_db import (
    upsert_product, insert_sales, insert_import, log_import,
    get_all_products, init_product_db
)

OHM_CALL_FILE = "OHM_call.xlsm"
KEHOACH_FILE = "OHM_KeHoachXoayVon.xlsx"


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def read_products_from_call():
    """Đọc sheet ThongTinSanPham — sản phẩm + giá bán, giá vốn, tồn kho."""
    if not os.path.exists(OHM_CALL_FILE):
        return [], f"File {OHM_CALL_FILE} not found"

    try:
        wb = openpyxl.load_workbook(OHM_CALL_FILE, data_only=True)
    except Exception as e:
        return [], f"Error opening file: {e}"

    products = []

    # Sheet: ThongTinSanPham
    if "ThongTinSanPham" not in wb.sheetnames:
        wb.close()
        return [], "Sheet 'ThongTinSanPham' not found"

    ws = wb["ThongTinSanPham"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        if not row or not row[3]:
            continue
        nhom1 = safe_str(row[0])
        nhom2 = safe_str(row[1])
        nhom3 = safe_str(row[2])
        ma_hang = safe_str(row[3])
        ten_hang = safe_str(row[4])
        thuong_hieu = safe_str(row[5])
        gia_ban = safe_float(row[6])
        gia_von = safe_float(row[7])
        ton_kho = safe_int(row[8])
        vi_tri = safe_str(row[9]) if len(row) > 9 else ""

        if not ma_hang:
            continue

        products.append({
            "ma_hang": ma_hang,
            "ten_hang": ten_hang,
            "thuong_hieu": thuong_hieu,
            "nhom_hang_cap1": nhom1,
            "nhom_hang_cap2": nhom2,
            "nhom_hang_cap3": nhom3,
            "gia_ban": gia_ban,
            "gia_von": gia_von,
            "ton_kho": ton_kho,
            "vi_tri": vi_tri,
        })

    wb.close()
    return products, None


def read_sales_from_call():
    """Đọc sheet LichSuBanHang — lịch sử bán hàng."""
    if not os.path.exists(OHM_CALL_FILE):
        return [], f"File {OHM_CALL_FILE} not found"

    try:
        wb = openpyxl.load_workbook(OHM_CALL_FILE, data_only=True)
    except Exception as e:
        return [], f"Error opening file: {e}"

    sales = []

    if "LichSuBanHang" not in wb.sheetnames:
        wb.close()
        return [], None

    ws = wb["LichSuBanHang"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        if not row or not row[1]:
            continue
        ma_hang = safe_str(row[1])
        if not ma_hang:
            continue

        thoi_gian = str(row[0]) if row[0] else datetime.now().isoformat()
        ten_hang = safe_str(row[2])
        sl_ban = safe_int(row[3])
        gia_ban = safe_float(row[4])
        thanh_tien = safe_float(row[5])
        gia_von = safe_float(row[6])
        loi_nhuan = safe_float(row[7])
        ten_khach = safe_str(row[8]) if len(row) > 8 else ""
        ma_kh = safe_str(row[9]) if len(row) > 9 else ""
        sl_tra = safe_int(row[10]) if len(row) > 10 else 0
        gt_tra = safe_float(row[11]) if len(row) > 11 else 0

        sales.append({
            "ma_hang": ma_hang,
            "ten_hang": ten_hang,
            "thoi_gian": thoi_gian,
            "sl_ban": sl_ban,
            "gia_ban": gia_ban,
            "thanh_tien": thanh_tien,
            "gia_von": gia_von,
            "loi_nhuan": loi_nhuan,
            "ten_khach_hang": ten_khach,
            "ma_kh": ma_kh,
            "sl_tra": sl_tra,
            "gt_tra": gt_tra,
        })

    wb.close()
    return sales, None


def read_imports_from_call():
    """Đọc sheet LichSuNhap — lịch sử nhập hàng."""
    if not os.path.exists(OHM_CALL_FILE):
        return [], f"File {OHM_CALL_FILE} not found"

    try:
        wb = openpyxl.load_workbook(OHM_CALL_FILE, data_only=True)
    except Exception as e:
        return [], f"Error opening file: {e}"

    imports = []

    if "LichSuNhap" not in wb.sheetnames:
        wb.close()
        return [], None

    ws = wb["LichSuNhap"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        if not row or not row[3]:
            continue
        ma_hang = safe_str(row[3])
        if not ma_hang:
            continue

        ngay_nhap = str(row[0]) if row[0] else datetime.now().isoformat()
        so_invoice = safe_str(row[1])
        ncc = safe_str(row[2])
        ten_hang = safe_str(row[4])
        so_luong = safe_int(row[5])
        gia_von_dv = safe_float(row[6])
        thanh_tien = safe_float(row[7])
        ghi_chu = safe_str(row[8]) if len(row) > 8 else ""

        imports.append({
            "ma_hang": ma_hang,
            "ten_hang": ten_hang,
            "ngay_nhap": ngay_nhap,
            "so_luong": so_luong,
            "gia_von_dv": gia_von_dv,
            "thanh_tien": thanh_tien,
            "nha_cung_cap": ncc,
            "so_invoice": so_invoice,
            "ghi_chu": ghi_chu,
        })

    wb.close()
    return imports, None


def import_all_from_call():
    """Import toàn bộ dữ liệu từ OHM_call.xlsm vào database."""
    products, err = read_products_from_call()
    if err:
        log_import("OHM_call.xlsm", 0, "error", err)
        return {"error": err}

    count = 0
    for p in products:
        upsert_product(p)
        count += 1

    sales, _ = read_sales_from_call()
    if sales:
        for s in sales:
            insert_sales(s)

    imports_data, _ = read_imports_from_call()
    if imports_data:
        for im in imports_data:
            insert_import(im)

    log_import("OHM_call.xlsm", count, "success", f"Imported {count} products, {len(sales or [])} sales, {len(imports_data or [])} imports")
    return {
        "products": count,
        "sales": len(sales or []),
        "imports": len(imports_data or []),
    }


def export_products_to_json():
    """Xuất danh sách sản phẩm ra JSON string."""
    products = get_all_products()
    return [
        {
            "ma_hang": p["ma_hang"],
            "ten_hang": p["ten_hang"],
            "nhom_hang": p["nhom_hang_cap1"],
            "gia_ban": p["gia_ban"],
            "gia_von": p["gia_von"],
            "ton_kho": p["ton_kho"],
            "gia_tri_ton": p["gia_von"] * p["ton_kho"],
            "bien_lai": round((p["gia_ban"] - p["gia_von"]) / p["gia_ban"] * 100, 2) if p["gia_ban"] else 0,
        }
        for p in products
    ]
